import json
import sys
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import unquote_plus

try:
    from common.config import get_env
    from common.dynamodb import mark_batch_failed, update_batch_processing
    from common.responses import json_response
except ModuleNotFoundError:
    sys.path.append(str(Path(__file__).resolve().parents[2] / "layers/common/python"))
    from common.config import get_env
    from common.dynamodb import mark_batch_failed, update_batch_processing
    from common.responses import json_response

EXPECTED_HEADERS = {
    "InvoiceNumber": "invoiceNumber",
    "Date": "date",
    "CustomerCUIT": "customerCUIT",
    "CustomerName": "customerName",
    "CustomerAddress": "customerAddress",
    "Amount": "amount",
    "TaxCode": "taxCode",
    "Description": "description",
    "Quantity": "quantity",
    "UnitPrice": "unitPrice",
    "IVA": "iva",
}

_S3_CLIENT = None
_SQS_CLIENT = None


def lambda_handler(event: dict[str, Any], context: Any) -> dict[str, Any]:
    summaries = []

    for record in event.get("Records", []):
        bucket, key = _s3_bucket_key(record)
        file_name = key.split("/")[-1]
        batch_id = _batch_id_from_key(key)

        if not batch_id:
            print(f"Ignorando objeto fuera del formato esperado: {key}")
            continue

        try:
            summary = process_object(bucket=bucket, key=key, batch_id=batch_id)
            summaries.append(summary)
            print(json.dumps(summary))
        except Exception as exc:
            message = f"Error tecnico en parser: {exc}"
            print(message)
            mark_batch_failed(batch_id, message, file_name=file_name, s3_key=key)
            raise

    return json_response(
        {
            "message": "Evento S3 procesado",
            "records": summaries,
            "count": len(summaries),
        }
    )


def process_object(*, bucket: str, key: str, batch_id: str) -> dict[str, Any]:
    file_name = key.split("/")[-1]
    workbook_bytes = _download_object(bucket, key)
    messages = parse_excel(workbook_bytes, batch_id=batch_id)

    if not messages:
        mark_batch_failed(
            batch_id,
            "El Excel no contiene facturas no vacias",
            file_name=file_name,
            s3_key=key,
        )
        return {
            "batchId": batch_id,
            "fileName": file_name,
            "totalInvoices": 0,
            "queuedInvoices": 0,
            "status": "FAILED",
        }

    update_batch_processing(
        batch_id=batch_id,
        file_name=file_name,
        s3_key=key,
        total_invoices=len(messages),
    )
    queued = _send_messages(messages)

    return {
        "batchId": batch_id,
        "fileName": file_name,
        "totalInvoices": len(messages),
        "queuedInvoices": queued,
        "status": "PROCESSING",
    }


def parse_excel(content: bytes, *, batch_id: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)

    if not headers:
        return []

    header_map = _header_map(headers)
    messages = []

    for row_number, row in enumerate(rows, start=2):
        if _is_empty_row(row):
            continue

        invoice = _invoice_from_row(row, header_map)
        invoice_number = invoice.get("invoiceNumber")
        if not invoice_number:
            invoice["invoiceNumber"] = f"ROW#{row_number}"

        messages.append(
            {
                "batchId": batch_id,
                "rowNumber": row_number,
                "invoice": invoice,
            }
        )

    return messages


def _header_map(headers: tuple[Any, ...]) -> dict[str, int]:
    normalized = {_normalize_header(header): index for index, header in enumerate(headers)}
    return {
        output_name: normalized[expected_header.lower()]
        for expected_header, output_name in EXPECTED_HEADERS.items()
        if expected_header.lower() in normalized
    }


def _invoice_from_row(row: tuple[Any, ...], header_map: dict[str, int]) -> dict[str, str]:
    invoice = {}
    for output_name, index in header_map.items():
        value = row[index] if index < len(row) else None
        invoice[output_name] = _normalize_cell(value, preserve_text=output_name == "invoiceNumber")
    return invoice


def _normalize_cell(value: Any, *, preserve_text: bool = False) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, Decimal):
        return format(value, "f")

    if isinstance(value, float):
        decimal_value = Decimal(str(value))
        return format(decimal_value, "f")

    if isinstance(value, int):
        return str(value)

    text = str(value).strip()
    if preserve_text:
        return text
    return text


def _normalize_header(header: Any) -> str:
    return "" if header is None else str(header).strip().lower()


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in row)


def _send_messages(messages: list[dict[str, Any]]) -> int:
    queue_url = get_env("QUEUE_URL", required=True)
    queued = 0

    for chunk in _chunks(messages, 10):
        entries = [
            {
                "Id": f"invoice-{message['rowNumber']}",
                "MessageBody": json.dumps(message),
            }
            for message in chunk
        ]
        response = _sqs_client().send_message_batch(QueueUrl=queue_url, Entries=entries)
        failed = response.get("Failed", [])
        if failed:
            raise RuntimeError(f"No se pudieron enviar mensajes a SQS: {failed}")
        queued += len(response.get("Successful", entries))

    return queued


def _chunks(items: list[dict[str, Any]], size: int):
    for index in range(0, len(items), size):
        yield items[index : index + size]


def _download_object(bucket: str, key: str) -> bytes:
    response = _s3_client().get_object(Bucket=bucket, Key=key)
    return response["Body"].read()


def _s3_bucket_key(record: dict[str, Any]) -> tuple[str, str]:
    s3_info = record.get("s3", {})
    bucket = s3_info.get("bucket", {}).get("name")
    key = unquote_plus(s3_info.get("object", {}).get("key", ""))

    if not bucket or not key:
        raise ValueError("Evento S3 sin bucket o key")

    return bucket, key


def _batch_id_from_key(key: str) -> str | None:
    parts = key.split("/")
    if len(parts) >= 3 and parts[0] == "uploads" and parts[1]:
        return parts[1]
    return None


def _s3_client():
    global _S3_CLIENT

    if _S3_CLIENT is None:
        import boto3

        _S3_CLIENT = boto3.client("s3")

    return _S3_CLIENT


def _sqs_client():
    global _SQS_CLIENT

    if _SQS_CLIENT is None:
        import boto3

        _SQS_CLIENT = boto3.client("sqs")

    return _SQS_CLIENT
